import os
import cv2
import face_recognition
import pandas as pd
from openpyxl import Workbook, load_workbook


# Conecta a webcam
webcam = cv2.VideoCapture(0)

def camera():
    df_funcionarios = pd.read_excel('/home/vyral/Vídeos/Face_recognizer/Funcionarios.xlsx')
    df_convidados = pd.read_excel('/home/vyral/Vídeos/Face_recognizer/Convidados.xlsx')

    wb = load_workbook('/home/vyral/Vídeos/Face_recognizer/teste.xlsx')
    ws = wb.active
    my_headder = []


    imagens_referencia = []

    # Procura em uma pasta específica o número de imagens, sempre nomeados com: rosto_salvo_i
    # Onde i = número da imagem, sempre em ordem crescente começando do 0
    for arquivo in os.listdir("/home/vyral/Vídeos/Face_recognizer/funcionarios"):

        # Verifica se o arquivo é uma imagem
        if arquivo.endswith(".jpg"):
            nome_sem_extensao = os.path.splitext(arquivo)[0]
            imagem_referencia = face_recognition.load_image_file(os.path.join("/home/vyral/Vídeos/Face_recognizer/funcionarios", arquivo))
            features_referencia = face_recognition.face_encodings(imagem_referencia)[0]
            imagens_referencia.append(features_referencia)

        imagens_referencia = [face_recognition.face_encodings(face_recognition.load_image_file(os.path.join("/home/vyral/Vídeos/Face_recognizer/funcionarios", f"{nome_sem_extensao}.jpg")))[0] for i in range(len(imagens_referencia))]
        if nome_sem_extensao not in df_funcionarios and nome_sem_extensao not in df_convidados:
        # Adicionar nome à planilha de convidados
            novo_registro = [{'Nome': nome_sem_extensao}]
            df_convidados = pd.concat([df_convidados, pd.DataFrame(novo_registro)], ignore_index=True)
            
    visitas = ''
    nao_reconhece = 0
    convidados = []
    while webcam.isOpened():
        validacao, frame = webcam.read()
        if not validacao:
            break

        # Converte a imagem da webcam para o padrão RGB
        imagem = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        cv2.imshow("Rosto", frame)

        # Encontra todos os rostos na imagem da webcam
        rostos = face_recognition.face_locations(imagem)

        # Extrai as features faciais de cada rosto encontrado
        features_rostos = face_recognition.face_encodings(imagem, rostos)

        # Verifica se algum dos rostos encontrados é conhecido
        rosto_reconhecido = False
        for features_rosto in features_rostos:
            for features_referencia in imagens_referencia:
                distancia = face_recognition.face_distance([features_referencia], features_rosto)
                if distancia < 0.5:
                    print("Rosto reconhecido")
                    rosto_reconhecido = True
                    break
            if rosto_reconhecido:
                break

        # Se nenhum rosto reconhecido foi encontrado, pergunte ao usuário se ele quer entrar como visitante
        if not rosto_reconhecido:
            nao_reconhece += 1
            if nao_reconhece == len(imagens_referencia):
                print("Rosto não reconhecido, deseja entrar como visitante?(S/N)")
                visitas = input().lower()
                if visitas == 's':
                    for i in range(nao_reconhece):
                        nome_convidado = input("Digite o nome completo do visitante: ")
                        convidados.append(nome_convidado)
                        my_headder.append('A'[i+1])
                        ws[i].fill = nome_convidado[i+1]
        if cv2.waitKey(1) == 27:  # ESC
            break

# Libera a memória
webcam.release()
cv2.destroyAllWindows()